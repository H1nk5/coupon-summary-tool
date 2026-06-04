import customtkinter as ctk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, date
from collections import defaultdict
import threading

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("代金券汇总工具")
        self.geometry("700x420")
        self.resizable(False, False)

        self.兰考路径 = ctk.StringVar()
        self.民权路径 = ctk.StringVar()
        self.输出路径 = ctk.StringVar()

        # === 标题栏 ===
        header = ctk.CTkFrame(self, fg_color="#1a73e8", corner_radius=0, height=50)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkLabel(
            header, text="代金券汇总工具",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="white",
        ).pack(expand=True)

        # === 主体 ===
        body = ctk.CTkFrame(self, fg_color="#f5f5f5", corner_radius=0)
        body.pack(fill="both", expand=True)

        inner = ctk.CTkFrame(body, fg_color="transparent")
        inner.pack(pady=25, padx=40, fill="x")

        # -- 兰考 --
        row1 = ctk.CTkFrame(inner, fg_color="transparent")
        row1.pack(fill="x", pady=(0, 12))
        ctk.CTkLabel(row1, text="兰考表格", font=ctk.CTkFont(size=13), width=80, anchor="e").pack(side="left")
        ctk.CTkEntry(
            row1, textvariable=self.兰考路径, height=38,
            font=ctk.CTkFont(size=12), placeholder_text="请选择兰考.xlsx",
        ).pack(side="left", fill="x", expand=True, padx=(10, 10))
        ctk.CTkButton(
            row1, text="浏览", width=90, height=38,
            font=ctk.CTkFont(size=13, weight="bold"),
            command=lambda: self.选择文件(self.兰考路径),
        ).pack(side="left")

        # -- 民权 --
        row2 = ctk.CTkFrame(inner, fg_color="transparent")
        row2.pack(fill="x", pady=(0, 12))
        ctk.CTkLabel(row2, text="民权表格", font=ctk.CTkFont(size=13), width=80, anchor="e").pack(side="left")
        ctk.CTkEntry(
            row2, textvariable=self.民权路径, height=38,
            font=ctk.CTkFont(size=12), placeholder_text="请选择民权.xlsx",
        ).pack(side="left", fill="x", expand=True, padx=(10, 10))
        ctk.CTkButton(
            row2, text="浏览", width=90, height=38,
            font=ctk.CTkFont(size=13, weight="bold"),
            command=lambda: self.选择文件(self.民权路径),
        ).pack(side="left")

        # -- 输出 --
        row3 = ctk.CTkFrame(inner, fg_color="transparent")
        row3.pack(fill="x", pady=(0, 20))
        ctk.CTkLabel(row3, text="输出位置", font=ctk.CTkFont(size=13), width=80, anchor="e").pack(side="left")
        ctk.CTkEntry(
            row3, textvariable=self.输出路径, height=38,
            font=ctk.CTkFont(size=12), placeholder_text="选择保存位置",
        ).pack(side="left", fill="x", expand=True, padx=(10, 10))
        ctk.CTkButton(
            row3, text="浏览", width=90, height=38,
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self.选择输出,
        ).pack(side="left")

        # -- 进度条 --
        self.进度 = ctk.CTkProgressBar(inner, height=14, corner_radius=7)
        self.进度.pack(fill="x", pady=(0, 6))
        self.进度.set(0)

        # -- 状态 --
        self.状态 = ctk.CTkLabel(inner, text="就绪", font=ctk.CTkFont(size=12), text_color="gray")
        self.状态.pack(pady=(0, 15))

        # -- 开始按钮 --
        self.按钮 = ctk.CTkButton(
            inner, text="开始汇总", height=48, width=220,
            font=ctk.CTkFont(size=16, weight="bold"),
            fg_color="#1a73e8", hover_color="#1558b0",
            command=self.开始,
        )
        self.按钮.pack()

    def 选择文件(self, var):
        path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx")])
        if path:
            var.set(path)

    def 选择输出(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile="代金券汇总表.xlsx",
        )
        if path:
            self.输出路径.set(path)

    def 更新状态(self, text, color="gray"):
        self.状态.configure(text=text, text_color=color)
        self.update_idletasks()

    def 开始(self):
        if not all([self.兰考路径.get(), self.民权路径.get(), self.输出路径.get()]):
            messagebox.showwarning("提示", "请先选择所有文件路径")
            return
        self.按钮.configure(state="disabled")
        threading.Thread(target=self.执行, daemon=True).start()

    def 读取数据(self, filepath):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            核销时间 = row[0]
            商品类型 = str(row[1]) if row[1] else ""
            订单实收 = float(row[2]) if row[2] else 0
            服务费 = float(row[3]) if row[3] else 0
            商家应得 = float(row[4]) if row[4] else 0
            if isinstance(核销时间, datetime):
                d = 核销时间.date()
            elif isinstance(核销时间, date):
                d = 核销时间
            else:
                d = datetime.strptime(str(核销时间)[:10], "%Y-%m-%d").date()
            is_代金券 = "代金券" in 商品类型
            data.append((d, is_代金券, 订单实收, 服务费, 商家应得))
        return data

    def 聚合(self, data):
        result = defaultdict(lambda: [0, 0, 0])
        for d, is_代金券, 实收, 服务费, 商家应得 in data:
            key = (d, is_代金券)
            result[key][0] += 实收
            result[key][1] += 服务费
            result[key][2] += 商家应得
        return result

    def 执行(self):
        try:
            self.更新状态("读取兰考数据...", "#1a73e8")
            self.进度.set(0.1)
            兰考 = self.读取数据(self.兰考路径.get())

            self.更新状态("读取民权数据...", "#1a73e8")
            self.进度.set(0.25)
            民权 = self.读取数据(self.民权路径.get())

            self.更新状态("汇总中...", "#1a73e8")
            self.进度.set(0.4)
            兰考_agg = self.聚合(兰考)
            民权_agg = self.聚合(民权)
            all_dates = sorted(set(k[0] for k in list(兰考_agg.keys()) + list(民权_agg.keys())))

            self.更新状态("生成表格...", "#1a73e8")
            self.进度.set(0.55)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "汇总"

            header_font = Font(name="等线", size=11, bold=True)
            header_align = Alignment(horizontal="center", vertical="center")
            data_font = Font(name="等线", size=11)
            data_align = Alignment(horizontal="right", vertical="center")
            thin_side = Side(style="thin")
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            # 颜色定义
            green_fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
            orange_fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")

            def set_cell(row, col, value, font=None, align=None, fill=None, border=None, num_fmt=None):
                cell = ws.cell(row=row, column=col, value=value)
                if font: cell.font = font
                if align: cell.alignment = align
                if fill: cell.fill = fill
                if border: cell.border = border
                if num_fmt: cell.number_format = num_fmt
                return cell

            ws.merge_cells("B1:J1")
            set_cell(1, 2, "商品", header_font, header_align)
            ws.merge_cells("K1:S1")
            set_cell(1, 11, "代金券", header_font, header_align)

            ws.merge_cells("A2:A3")
            set_cell(2, 1, "日期", header_font, header_align, border=thin_border)
            set_cell(3, 1, None, border=thin_border)

            # Row 2: 合计/兰考/民权 with colors
            groups = [
                ("B2:D2", "合计", green_fill), ("E2:G2", "兰考", yellow_fill), ("H2:J2", "民权", orange_fill),
                ("K2:M2", "合计", green_fill), ("N2:P2", "兰考", yellow_fill), ("Q2:S2", "民权", orange_fill),
            ]
            for merge_range, label, fill in groups:
                ws.merge_cells(merge_range)
                col = openpyxl.utils.cell.column_index_from_string(merge_range[0])
                set_cell(2, col, label, header_font, header_align, fill)

            # Row 3: column headers with matching colors
            col3_fills = [green_fill]*3 + [yellow_fill]*3 + [orange_fill]*3 + [green_fill]*3 + [yellow_fill]*3 + [orange_fill]*3
            cols3 = ["订单实收", "服务费", "商家应得"] * 6
            for j, (label, fill) in enumerate(zip(cols3, col3_fills)):
                set_cell(3, 2 + j, label, header_font, header_align, fill)

            # Borders for header rows
            for r in range(1, 4):
                for c in range(1, 20):
                    ws.cell(row=r, column=c).border = thin_border

            # Column widths
            ws.column_dimensions["A"].width = 8.3
            ws.column_dimensions["B"].width = 16.3
            ws.column_dimensions["K"].width = 16.3
            for col in "CDEFGHIJLMNOPQRS":
                ws.column_dimensions[col].width = 13

            # Data rows
            total = len(all_dates)
            for i, d in enumerate(all_dates):
                row = 4 + i
                cell = ws.cell(row=row, column=1, value=datetime(d.year, d.month, d.day))
                cell.number_format = "M/D"
                cell.font = header_font
                cell.alignment = header_align
                cell.border = thin_border

                lk_g = 兰考_agg.get((d, False), [0, 0, 0])
                mq_g = 民权_agg.get((d, False), [0, 0, 0])
                tot_g = [lk_g[j] + mq_g[j] for j in range(3)]

                lk_c = 兰考_agg.get((d, True), [0, 0, 0])
                mq_c = 民权_agg.get((d, True), [0, 0, 0])
                tot_c = [lk_c[j] + mq_c[j] for j in range(3)]

                all_vals = tot_g + lk_g + mq_g + tot_c + lk_c + mq_c
                for j, val in enumerate(all_vals):
                    c = ws.cell(row=row, column=2 + j, value=round(val, 2))
                    c.font = data_font
                    c.alignment = data_align
                    c.border = thin_border
                    c.number_format = "#,##0.00"

                self.进度.set(0.55 + 0.4 * (i + 1) / total)

            self.更新状态("保存中...", "#1a73e8")
            self.进度.set(0.95)
            wb.save(self.输出路径.get())

            self.进度.set(1.0)
            self.更新状态(f"完成! {total} 天 | {all_dates[0]} ~ {all_dates[-1]}", "#16a34a")
            messagebox.showinfo("完成", f"汇总完成!\n共 {total} 天\n{all_dates[0]} ~ {all_dates[-1]}\n\n已保存: {self.输出路径.get()}")

        except Exception as e:
            self.更新状态(f"出错: {e}", "#dc2626")
            messagebox.showerror("错误", str(e))
        finally:
            self.按钮.configure(state="normal")


if __name__ == "__main__":
    app = App()
    app.mainloop()
